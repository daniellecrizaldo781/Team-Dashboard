import openpyxl
w = openpyxl.load_workbook("sched.xlsx", data_only=True)
s = w["Leave Request Sheet"]
rows = list(s.iter_rows(values_only=True))
print("dims:", s.max_row, "x", s.max_column)
for i, r in enumerate(rows[:6]):
    print(i+1, [str(c)[:26] if c is not None else None for c in r[:12]])
print("\n--- last populated rows ---")
pop = [i for i, r in enumerate(rows) if any(c is not None and str(c).strip() for c in r)]
print("last populated row index (1-based):", pop[-1]+1, "| count:", len(pop))
for i in pop[-3:]:
    print(i+1, [str(c)[:26] if c is not None else None for c in rows[i][:12]])
print("\n--- distinct values per col ---")
hdr = [str(c).strip() if c else '' for c in rows[0][:12]]
for ci, h in enumerate(hdr):
    vals = {str(rows[i][ci]).strip() for i in pop[1:] if rows[i][ci] is not None}
    if h and len(vals) <= 8:
        print(f"  col{ci} {h!r}: {sorted(vals)[:8]}")
    elif h:
        print(f"  col{ci} {h!r}: {len(vals)} distinct")
