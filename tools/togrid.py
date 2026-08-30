import openpyxl, json, datetime, os, re, zipfile
import xml.etree.ElementTree as ET

NS = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'

# Hidden tabs are deliberately excluded: they are drafts, archives and old
# copies that the team has taken out of circulation. Filtering here means every
# downstream parser only ever sees live tabs, with no per-parser opt-out.


def parse_shared_strings_runs(z):
    """Return list (index -> list of (text, bold, italic)) for rich shared strings."""
    out = []
    try:
        data = z.read('xl/sharedStrings.xml').decode('utf-8', 'ignore')
    except KeyError:
        return out
    root = ET.fromstring(data)
    for si in root.findall(NS + 'si'):
        runs = []
        # A shared string is either a single <t> or multiple <r> runs.
        t = si.find(NS + 't')
        if t is not None and len(si.findall(NS + 'r')) == 0:
            runs.append((t.text or '', False, False))
        else:
            for r in si.findall(NS + 'r'):
                rpr = r.find(NS + 'rPr')
                bold = False
                italic = False
                if rpr is not None:
                    if rpr.find(NS + 'b') is not None:
                        # <b val="0"> means NOT bold; <b/> or <b val="1"> means bold
                        b = rpr.find(NS + 'b')
                        bold = (b.get('val') in (None, '1', 'true'))
                    if rpr.find(NS + 'i') is not None:
                        i = rpr.find(NS + 'i')
                        italic = (i.get('val') in (None, '1', 'true'))
                tnode = r.find(NS + 't')
                runs.append((tnode.text if tnode is not None else '', bold, italic))
        out.append(runs)
    return out


def cell_runs(z, sheet_xml, sst_runs):
    """Build a map (row,col) -> shared-string-index for cells with t='s'."""
    # We only need indices for cells; openpyxl already gives us text via value,
    # so we map by reading the raw <c> elements to find the shared-string index.
    data = z.read(sheet_xml).decode('utf-8', 'ignore')
    idx = {}
    for m in re.finditer(r'<c r="([A-Z]+)(\d+)"[^>]*t="s"[^>]*>(?:<f[^>]*>.*?</f>)?<v>(\d+)</v></c>', data):
        col, row, sidx = m.group(1), int(m.group(2)), int(m.group(3))
        idx[(row, col_letter_to_num(col))] = sidx
    return idx


def col_letter_to_num(letters):
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch.upper()) - ord('A') + 1)
    return n


def grid(path, out):
    z = zipfile.ZipFile(path)
    sst_runs = parse_shared_strings_runs(z)
    w = openpyxl.load_workbook(path, data_only=True)
    d = {}
    skipped = []
    for n in w.sheetnames:
        s = w[n]
        if s.sheet_state != "visible":
            skipped.append(n)
            continue
        # find sheet xml name for this sheet to read rich runs
        # openpyxl sheet index maps to sheetN.xml in document order
        sheet_idx = w.sheetnames.index(n) + 1
        sst_map = {}
        try:
            sst_map = cell_runs(z, 'xl/worksheets/sheet%d.xml' % sheet_idx, sst_runs)
        except KeyError:
            pass
        rows = []
        for r in s.iter_rows(min_row=1, max_row=s.max_row or 1):
            row = []
            for c in r:
                val = None
                if c.value is None:
                    val = ''
                elif isinstance(c.value, datetime.datetime):
                    val = {"__d": c.value.strftime("%Y-%m-%d")}
                elif isinstance(c.value, datetime.date):
                    val = {"__d": c.value.strftime("%Y-%m-%d")}
                elif isinstance(c.value, datetime.time):
                    val = c.value.strftime("%H:%M:%S")
                elif isinstance(c.value, datetime.timedelta):
                    val = str(c.value)
                else:
                    # possible rich text?
                    key = (c.row, col_letter_to_num(c.column_letter if isinstance(c.column_letter, str) else c.column))
                    sidx = sst_map.get(key)
                    if sidx is not None and sidx < len(sst_runs) and len(sst_runs[sidx]) > 1:
                        # rich text -> structured runs
                        runs = [list(x) for x in sst_runs[sidx]]
                        val = {"__rt": runs}
                    else:
                        val = c.value
                row.append(val)
            rows.append(row)
        d[n] = rows
    json.dump(d, open(out, "w"), ensure_ascii=False)
    print(out, len(d), "visible tabs")
    if skipped:
        print("   skipped %d hidden: %s" % (len(skipped), ", ".join(skipped)))


grid("qa.xlsx", "qa.json")
grid("sched.xlsx", "sched.json")
if os.path.exists("casc.xlsx"):
    grid("casc.xlsx", "casc.json")
if os.path.exists("prod.xlsx"):
    grid("prod.xlsx", "prod.json")
