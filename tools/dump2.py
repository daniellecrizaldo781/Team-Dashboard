import openpyxl
w = openpyxl.load_workbook("sched.xlsx", data_only=True)
for n in ["Break Schedule", "Leave Request Sheet", "OT SCHEDULE",
          "NEW TEAM BREAK SCHEDULE", "DAILY SBS SCHEDULE",
          "BREAK SCHEDULE AND ADMIN TASK A"]:
    s = w[n]
    print("=== TAB:", n, s.max_row, "x", s.max_column)
    for r in s.iter_rows(min_row=1, max_row=min(8, s.max_row),
                         max_col=min(s.max_column, 18), values_only=True):
        c = [str(x)[:20] if x is not None else "" for x in r]
        while c and c[-1] == "":
            c.pop()
        if c:
            print("  |".join(c))
    print()
