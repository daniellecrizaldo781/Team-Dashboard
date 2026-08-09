import openpyxl, sys

def dump(f, tabs, label):
    w = openpyxl.load_workbook(f, data_only=True)
    print("#####", label)
    for n in tabs:
        if n not in w.sheetnames:
            print("MISSING TAB", n); continue
        s = w[n]
        print("=== TAB:", n, s.max_row, "x", s.max_column)
        for r in s.iter_rows(min_row=1, max_row=min(7, s.max_row),
                             max_col=min(s.max_column, 24), values_only=True):
            cells = [str(c)[:22] if c is not None else "" for c in r]
            while cells and cells[-1] == "":
                cells.pop()
            if cells:
                print("  |".join(cells))
        print()

dump("qa.xlsx", ["Team Info", "Team Weekly and Monthly Stats", "TEAM STATS",
                 "Daily and Weekly Call Stats", "WEEKLY SCORECARD", "Dan",
                 "Cherry", "MONTHLY SCORECARD", "OHA Call Callibration"], "QA")
dump("sched.xlsx", ["Team Schedule", "Break Schedule", "Leave Request Sheet",
                    "OT SCHEDULE", "NEW TEAM BREAK SCHEDULE",
                    "DAILY SBS SCHEDULE"], "SCHED")
